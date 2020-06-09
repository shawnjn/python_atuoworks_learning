#时间：2020.05.30
#环境：python 3.7 + mac
#编写:shawn_jn
#主要内容：打开指定库存文件excel，选择需要操作的sheet；
# 对入库数据进行增加，若增加物资已存在，提醒已存在，并显示存储单位，在sheet中新建行存储；提取当前时间

import os
import time
import datetime
import openpyxl as xl

#获取文件修改时间，在保存文件时增加时间；【需放至最后，该段程序未完成】
time_mark = os.stat('库存管理-gxj.xlsx').st_mtime
time_tep = datetime.datetime.fromtimestamp(time_mark)
time_nm = time_tep.strftime('%Y-%m-%d-%H:%M')
# print(time_nm )

#表格字体等格式，时间等要求
timenow = datetime.datetime.now().strftime('%Y-%m-%d-%H:%M')
font_01 = Font(name='楷体-简', size=12, bold=True, italic=0, color="000000") #字体设置：bold粗体，italic斜体，黑色
font_02 = Font(name='楷体-简', size=12, bold=True, italic=0, color="FF0000") #字体设置：bold粗体，italic斜体，红色
alignment_01 = Alignment(horizontal='center', vertical='center',
                         text_rotation=0, wrap_text=True)   #对齐设置：text_rotation旋转角度,wrap_text=True自动换行
side_01 = Side(style='thin', color='000000')   #style=边线样式，color=边线颜色
border_01 = Border(left=side_01, right=side_01, top=side_01, bottom=side_01)    #设置边框样式：
fill_01 = PatternFill(fill_type='solid', fgColor='FFFF00')    #设置填充样式:fill_type=填充样式， fgColor=填充颜色 黄色
h_01 = 36   #行高设置

#搜索选择指定的excel文件（通过比较最近修改确定最新文件）
files_tep = []
files_stmtime =[]
for file in os.scandir():
    if '库存管理' in file.name:
        # print(file_num)
        files_tep.append(file.name)
        files_stmtime.append(os.stat(file))
    time_last_index = files_stmtime.index(max(files_stmtime))   #获取最近时间的文件index序号
# print(time_last_index)
# print(files[time_last_index])
file = files_tep[time_last_index]   #通过对应的序号，获取文件

#打开最新excel文件后,选择对应sheet
wb = xl.load_workbook(filename=file)
print(f'已打开  【{file}】  文件\nsheet列表如下：')
# print(wb.sheetnames, '\n请选择需要修改的文件：')
i = 1
wb_sheet_list = []
for sheetname in wb.sheetnames:
    print(str(i) + '-' + sheetname)
    wb_sheet_list.append(sheetname)
    i = i + 1
m = int(input('请选择sheet表格，输入数字：'))
ws2 = wb[wb_sheet_list[m-1]]
print(f'已激活"{wb_sheet_list[m-1]}"列表')

#【库存数据】按照格式输出库存sheet表格；
if m == 2:
    wb_tep1 = xl.Workbook()
    ws_tep1 = wb_tep1.active
    ws_tep1.title = f'库存查询'
    #复制数据至新表格，同时自定义表格格式
    for i in range(1, ws_x.max_row+1):
        for j in range(1, ws_x.max_column):
            ws_tep1.cell(row=i, column=j).value = ws_x.cell(row=i, column=j).value  #复制数据至新表格；
            ws_tep1.cell(row=i, column=j).border = border_01
            if i < 3:
                ws_tep1.cell(row=i, column=j).font = font_01
                ws_tep1.cell(row=i, column=j).alignment = alignment_01
                if j > 9 and j < 14:
                    ws_tep1.cell(row=i, column=j).font = font_02
    ws_tep1.cell(row=2, column=3).fill = fill_01
    ws_tep1.cell(row=2, column=10).fill = fill_01
    #定义行高
    ws_tep1.row_dimensions[1].height = h_01
    #自定义各列的宽度
    widths = [8, 8, 9, 12, 9, 9, 9, 8, 9, 9, 9, 15, 15] #自定义各列的宽度（特殊）
    column_tep = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    c_num = 0
    for w_num in widths:
        ws_tep1.column_dimensions[f'{column_tep[c_num]}'].width = w_num + 2
        # ws.column_dimensions['B'].width = 20
        c_num += 1
    ws_tep1.merge_cells('A1:M1')
    #判断库存数量是否低于安全库存量，接近时预警（黄色框）、低于报警（红色框）；
    for r in range(3, ws_tep1.max_row+1):
      num = ws_tep1.cell(row=r, column=8).value #库存数量
      num_war = ws_tep1.cell(row=r, column=10).value    #报警数量（定额）
      num_alarm = ws_tep1.cell(row=r, column=11).value    #预警（提醒）
      try:
          num_1 = num - num_alarm
          num_2 = num - num_war
          if num_1 <= 0:
              ws_tep1.cell(row=r, column=8).fill = fill_yel
              print(f'{ws_tep1.cell(row=r, column=5).value}-'
                    f'{ws_tep1.cell(row=r, column=6).value}------【预警】注意补充！')
          if num_2 <= 0:
              ws_tep1.cell(row=r, column=8).fill = fill_red
              print(f'\033[32;0m {ws_tep1.cell(row=r, column=5).value}'
                    f'-{ws_tep1.cell(row=r, column=6).value}'
                    f'---【报警、报警、报警】库存不足，请立即补充！！')
      except TypeError:
          print('【存在 None 单元格】')
          pass
    wb_tep1.save(f'库存查询结果-{timenow}.xlsx')
    print(f'\n------库存查询结束！------\n请查看《库存查询结果-{timenow}.xlsx》\n')

#【入库操作】在sheet内进行循环多次操作；（入库数据完成，剩余将入库数据汇总至库存数据，部分输入错误或格式要求未细化，表格格式未美化）
if m == 3:
    running = True
    ls_nm = 1
    while running:
        name1 = input('输入货品名称：')
        #获取目前库存信息，主要是便于入库是重复输入，相同物品即可进行自动输入，不同物品进行逐项输入
        ws1 = wb[wb_sheet_list[1]]
        row1 = ws1.max_row
        column1 = ws1.max_column
        row2 = ws2.max_row
        column2 = ws2.max_column
        # print(row, column)
        list_things = []
        for i in range(3, row1):
            value = ws1.cell(row=i, column=5).value
            # print(value)
            list_things.append(value)
        # 如果入库物品存在于库存中，将直接提取部分信息
        if name1 in list_things:
            index1 = list_things.index(name1) + 3
            print(list_things)
            print(index1)
            mark1 = ws1.cell(row=index1, column=2).value  # 获取输入物品'物料编码'
            mark2 = ws1.cell(row=index1, column=3).value  # 获取输入物品'唯一识别码'
            mark3 = ws1.cell(row=index1, column=4).value  # 获取输入物品'货物类别'
            mark4 = ws1.cell(row=index1, column=6).value  # 获取输入物品'规格型号'
            mark5 = ws1.cell(row=index1, column=7).value  # 获取输入物品'单位'
            mark6 = ws1.cell(row=index1, column=12).value  # 获取输入物品'是否电动、起重、安全工器具'
            #将以上信息自动写入"入库数据"中
            ws2.cell(row=row2 + ls_nm, column=2).value = datetime.datetime.now().strftime('%Y-%m-%d-%H:%M') #时间输入
            ws2.cell(row=row2 + ls_nm, column=3).value = mark1
            ws2.cell(row=row2 + ls_nm, column=4).value = mark2
            ws2.cell(row=row2 + ls_nm, column=5).value = mark3
            ws2.cell(row=row2 + ls_nm, column=6).value = name1
            ws2.cell(row=row2 + ls_nm, column=7).value = mark4
            ws2.cell(row=row2 + ls_nm, column=8).value = mark5
            ws2.cell(row=row2 + ls_nm, column=12).value = mark6

            name2 = input('输入入库数量：')
            name3 = input('输入货品单价（元）：')
            name4 = input('输入经手人：')
            name5 = input('输入货品备注说明：')

            ws2.cell(row=row2 + ls_nm, column=9).value = name2
            ws2.cell(row=row2 + ls_nm, column=10).value = name3
            ws2.cell(row=row2 + ls_nm, column=11).value = name4
            ws2.cell(row=row2 + ls_nm, column=13).value = name5
        #如果入库物品不曾在库存，将新增
        else:
            print(f'"{name1}"库存内无该物品！')
            for i in range(1, column2):
                print(f"请输入'{ws2.cell(row=2, column=i + 2).value}'")
                text_in = input()
                ws2.cell(row=row2 + 1, column=i + 2).value = text_in
        q = input('是否需要继续入库操作：y/n')
        if q == 'n':
            running = False

#【出库操作】按照格式输出库存sheet表格；
if m == 4:
    1

#【库存数据】按照格式输出库存sheet表格；
if m == 5:
    1

#【库存数据】按照格式输出库存sheet表格；
if m == 6:
    1

#保存文件（文件开头增加修改完成时间，精确到分钟）
save_ch = 0
while save_ch == 0:
    save_done = input('是否需要对库存excel表格保存？'
                      '\n---保存：请输入"y/Y"'
                      '\n---不保存请输入"n/N"\n')
    if save_done == 'y' or save_done == 'Y':
        time_now_ct = datetime.datetime.now()
        time_now = time_now_ct.strftime('%Y-%m-%d-%H:%M')
        file_n = f"库存管理(修改时间：{str(time_now)})"
        wb.save(f'{file_n}.xlsx')
        print(f'操作完成！\n已保存为:{file_n}.xlsx文件 \n感谢使用！！！"')
        save_ch = 1
    elif save_done == 'n' or save_done == 'N':
        print('此次操作，库存excel未保存，感谢使用！！！')
        save_ch = 1
    else:
        print('输入错误，请重新输入！！！')
        continue

